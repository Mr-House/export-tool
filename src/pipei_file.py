# uncompyle6 version 3.7.4
# Python bytecode 3.7 (3394)
# Decompiled from: Python 3.7.3 (v3.7.3:ef4ec6ed12, Mar 25 2019, 22:22:05) [MSC v.1916 64 bit (AMD64)]
# Embedded file name: pipei_file.py
import pandas, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from set_time_path import get_Path
from main_ui import test


def print_ui(id, pp):
    for row in pp.itertuples():
        a = ''
        for i in range(1, len(row)):
            a = a + str(row[i]) + ' '

        test(a + '\n')


def pipei(data0, pipei_path):
    connectTable = pandas.read_excel('connectTable.xlsx')
    pipei_list = str(pipei_path).split(',')
    a = len(pipei_list)
    for i in range(0, int(int(a) / 3)):
        data0 = connectTable.loc[:, [pipei_list[(i + int(int(a) / 3))], pipei_list[(i + 2 * int(int(a) / 3))]]].merge(
            data0, left_on=(pipei_list[(i + int(int(a) / 3))]), right_on=(pipei_list[i]), how='right')
        if pipei_list[(i + int(int(a) / 3))] != pipei_list[i]:
            data0.drop([pipei_list[(i + int(int(a) / 3))]], axis=1, inplace=True)

    return data0


def to_excel(data, name):
    cwd = get_Path()
    wb = openpyxl.Workbook()
    ws_paigong = wb.create_sheet(title=name)
    rows_paigong = dataframe_to_rows(data, index=False)
    for r_idx, row in enumerate(rows_paigong, 1):
        for c_idx, value in enumerate(row, 1):
            ws_paigong.cell(row=r_idx, column=c_idx, value=value)

    wb.save(cwd + '/' + name + '.xlsx')

# okay decompiling .\pipei_file.pyc
